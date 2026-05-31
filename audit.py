#!/usr/bin/env python3
"""Expense audit script: downloads an .xlsx from Google Drive, audits expenses,
writes an 'Audit Report' sheet, and re-uploads the file."""

import argparse
import json
import re
import subprocess
import sys
import tempfile
import os
from collections import defaultdict
from datetime import datetime
from statistics import mean, stdev

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── column name mappings ──────────────────────────────────────────────────────
COLUMN_ALIASES = {
    "date":        {"дата", "date", "datum", "fecha", "data"},
    "amount":      {"сумма", "amount", "стоимость", "цена", "cost", "price",
                    "value", "расход", "итого", "total"},
    "category":    {"категория", "category", "тип", "type", "раздел", "section"},
    "description": {"описание", "description", "комментарий", "note", "memo",
                    "назначение", "детали", "details", "наименование"},
}


def extract_file_id(url: str) -> str:
    patterns = [
        r"/file/d/([a-zA-Z0-9_-]+)",
        r"/spreadsheets/d/([a-zA-Z0-9_-]+)",
        r"[?&]id=([a-zA-Z0-9_-]+)",
        r"^([a-zA-Z0-9_-]{20,})$",
    ]
    for pat in patterns:
        m = re.search(pat, url)
        if m:
            return m.group(1)
    sys.exit(f"[ERROR] Cannot extract file ID from URL: {url}")


def gws(*args) -> dict:
    cmd = ["gws"] + list(args)
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        sys.exit(f"[ERROR] gws command failed:\n{result.stderr}")
    if result.stdout.strip():
        return json.loads(result.stdout)
    return {}


def download_file(file_id: str, dest_path: str):
    cmd = [
        "gws", "drive", "files", "get",
        "--params", json.dumps({"fileId": file_id, "alt": "media"}),
        "--output", dest_path,
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        sys.exit(f"[ERROR] Download failed:\n{result.stderr}")


def upload_file(file_id: str, local_path: str):
    cmd = [
        "gws", "drive", "files", "update",
        "--params", json.dumps({"fileId": file_id}),
        "--upload", local_path,
        "--upload-content-type",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        sys.exit(f"[ERROR] Upload failed:\n{result.stderr}")


def detect_columns(headers: list) -> dict:
    """Map canonical names → 0-based column indices."""
    mapping = {}
    for idx, raw in enumerate(headers):
        if raw is None:
            continue
        normalized = str(raw).strip().lower()
        for canonical, aliases in COLUMN_ALIASES.items():
            if normalized in aliases and canonical not in mapping:
                mapping[canonical] = idx
    return mapping


def parse_amount(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^\d.,-]", "", str(value)).replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if hasattr(value, "timetuple"):
        return datetime(value.year, value.month, value.day)
    s = str(value).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def load_rows(ws, col_map: dict) -> list[dict]:
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        record = {
            "date":        parse_date(row[col_map["date"]] if "date" in col_map else None),
            "amount":      parse_amount(row[col_map["amount"]] if "amount" in col_map else None),
            "category":    str(row[col_map["category"]]).strip() if "category" in col_map and row[col_map["category"]] is not None else "Без категории",
            "description": str(row[col_map["description"]]).strip() if "description" in col_map and row[col_map["description"]] is not None else "",
        }
        rows.append(record)
    return rows


# ── styling helpers ───────────────────────────────────────────────────────────
def _header_style(cell, bg="1F4E79"):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def _section_title(ws, row, title):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=12, color="1F4E79")
    cell.fill = PatternFill("solid", fgColor="D6E4F0")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return row + 1


def _write_table(ws, start_row, headers, data_rows, warn_rows=None):
    warn_rows = warn_rows or set()
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c, value=h)
        _header_style(cell, bg="2E75B6")
        cell.border = _thin_border()

    for r_idx, row in enumerate(data_rows):
        excel_row = start_row + 1 + r_idx
        is_warn = r_idx in warn_rows
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if is_warn:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
        # zebra
        if not is_warn and r_idx % 2 == 1:
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=excel_row, column=c_idx).fill = PatternFill("solid", fgColor="EBF3FB")

    return start_row + 1 + len(data_rows) + 1


# ── audit logic ───────────────────────────────────────────────────────────────
def audit(rows: list[dict]) -> dict:
    valid = [r for r in rows if r["amount"] is not None]

    # 1. Totals by category
    cat_totals: dict[str, float] = defaultdict(float)
    cat_counts: dict[str, int] = defaultdict(int)
    cat_amounts: dict[str, list[float]] = defaultdict(list)
    for r in valid:
        cat_totals[r["category"]] += r["amount"]
        cat_counts[r["category"]] += 1
        cat_amounts[r["category"]].append(r["amount"])

    # 2. Top-10
    top10 = sorted(valid, key=lambda r: r["amount"], reverse=True)[:10]

    # 3. Anomalies: amount > 2× category mean
    anomalies = []
    for r in valid:
        cat = r["category"]
        if len(cat_amounts[cat]) >= 2:
            avg = mean(cat_amounts[cat])
            if r["amount"] > 2 * avg:
                anomalies.append({**r, "category_avg": avg})

    # 4. Duplicates: same date + amount + category
    seen: dict[tuple, list[int]] = defaultdict(list)
    for idx, r in enumerate(valid):
        key = (
            r["date"].date() if r["date"] else None,
            r["amount"],
            r["category"],
        )
        seen[key].append(idx)
    duplicates = [(k, idxs) for k, idxs in seen.items() if len(idxs) > 1]

    # 5. Monthly trend
    monthly: dict[str, float] = defaultdict(float)
    for r in valid:
        if r["date"]:
            key = r["date"].strftime("%Y-%m")
            monthly[key] += r["amount"]
    monthly_sorted = sorted(monthly.items())

    return {
        "total_rows": len(rows),
        "valid_rows": len(valid),
        "grand_total": sum(r["amount"] for r in valid),
        "cat_totals": cat_totals,
        "cat_counts": cat_counts,
        "cat_amounts": cat_amounts,
        "top10": top10,
        "anomalies": anomalies,
        "duplicates": duplicates,
        "monthly": monthly_sorted,
    }


def write_audit_sheet(wb: openpyxl.Workbook, result: dict, generated_at: datetime):
    sheet_name = "Audit Report"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # set column widths
    widths = [28, 16, 10, 28, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    cur_row = 1

    # ── Title ─────────────────────────────────────────────────────────────────
    title_cell = ws.cell(row=cur_row, column=1,
                         value="📊 Audit Report — Expense Analysis")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)
    ws.row_dimensions[cur_row].height = 28
    cur_row += 1

    meta_cell = ws.cell(row=cur_row, column=1,
                        value=f"Generated: {generated_at.strftime('%Y-%m-%d %H:%M')}  |  "
                              f"Rows read: {result['total_rows']}  |  "
                              f"Valid rows: {result['valid_rows']}")
    meta_cell.font = Font(italic=True, color="555555")
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)
    cur_row += 2

    # ── 1. Totals by category ─────────────────────────────────────────────────
    cur_row = _section_title(ws, cur_row, "1. Итого по категориям")
    headers = ["Категория", "Сумма", "Кол-во", "Ср. чек", "% от итого"]
    grand = result["grand_total"] or 1
    data = []
    for cat, total in sorted(result["cat_totals"].items(), key=lambda x: -x[1]):
        cnt = result["cat_counts"][cat]
        data.append([
            cat,
            round(total, 2),
            cnt,
            round(total / cnt, 2) if cnt else 0,
            f"{100 * total / grand:.1f}%",
        ])
    cur_row = _write_table(ws, cur_row, headers, data)

    # ── 2. Top-10 largest ─────────────────────────────────────────────────────
    cur_row = _section_title(ws, cur_row, "2. Топ-10 крупнейших трат")
    headers = ["#", "Дата", "Сумма", "Категория", "Описание"]
    data = []
    for i, r in enumerate(result["top10"], 1):
        data.append([
            i,
            r["date"].strftime("%d.%m.%Y") if r["date"] else "—",
            round(r["amount"], 2),
            r["category"],
            r["description"][:60] if r["description"] else "—",
        ])
    cur_row = _write_table(ws, cur_row, headers, data)

    # ── 3. Anomalies ──────────────────────────────────────────────────────────
    cur_row = _section_title(ws, cur_row, "3. Аномалии ⚠️  (сумма > 2× среднего по категории)")
    if result["anomalies"]:
        headers = ["⚠️", "Дата", "Сумма", "Категория", "Ср. по категории"]
        data = []
        for r in sorted(result["anomalies"], key=lambda x: -x["amount"]):
            data.append([
                "⚠️",
                r["date"].strftime("%d.%m.%Y") if r["date"] else "—",
                round(r["amount"], 2),
                r["category"],
                round(r["category_avg"], 2),
            ])
        warn_rows = set(range(len(data)))
        cur_row = _write_table(ws, cur_row, headers, data, warn_rows=warn_rows)
    else:
        ws.cell(row=cur_row, column=1, value="Аномалий не обнаружено").font = Font(italic=True, color="008000")
        cur_row += 2

    # ── 4. Duplicates ─────────────────────────────────────────────────────────
    cur_row = _section_title(ws, cur_row, "4. Дубликаты (дата + сумма + категория)")
    if result["duplicates"]:
        headers = ["Дата", "Сумма", "Категория", "Кол-во вхождений", ""]
        data = []
        for (dt, amt, cat), idxs in result["duplicates"]:
            data.append([
                dt.strftime("%d.%m.%Y") if dt else "—",
                round(amt, 2) if amt else "—",
                cat,
                len(idxs),
                "",
            ])
        cur_row = _write_table(ws, cur_row, headers, data,
                               warn_rows=set(range(len(data))))
    else:
        ws.cell(row=cur_row, column=1, value="Дубликатов не обнаружено").font = Font(italic=True, color="008000")
        cur_row += 2

    # ── 5. Monthly trend ──────────────────────────────────────────────────────
    cur_row = _section_title(ws, cur_row, "5. Тренд по месяцам")
    if result["monthly"]:
        headers = ["Месяц", "Сумма", "Δ к пред. месяцу", "", ""]
        data = []
        prev = None
        for month_str, total in result["monthly"]:
            try:
                label = datetime.strptime(month_str, "%Y-%m").strftime("%B %Y")
            except Exception:
                label = month_str
            delta = ""
            if prev is not None:
                diff = total - prev
                arrow = "▲" if diff > 0 else "▼"
                delta = f"{arrow} {abs(diff):,.2f} ({100*diff/prev:+.1f}%)" if prev else "—"
            data.append([label, round(total, 2), delta, "", ""])
            prev = total
        cur_row = _write_table(ws, cur_row, headers, data)
    else:
        ws.cell(row=cur_row, column=1, value="Нет данных с датами").font = Font(italic=True, color="888888")
        cur_row += 2

    # freeze top rows
    ws.freeze_panes = "A4"


# ── main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Expense audit — Google Drive .xlsx")
    parser.add_argument("--file-url", required=True, help="Google Drive file URL or ID")
    args = parser.parse_args()

    file_id = extract_file_id(args.file_url)
    print(f"[*] File ID: {file_id}")

    with tempfile.TemporaryDirectory() as tmpdir:
        local_path = os.path.join(tmpdir, "expenses.xlsx")

        print("[*] Downloading file from Google Drive…")
        download_file(file_id, local_path)

        print("[*] Reading workbook…")
        wb = openpyxl.load_workbook(local_path)
        ws = wb.worksheets[0]
        print(f"    Sheet: '{ws.title}', dims: {ws.dimensions}")

        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col_map = detect_columns(headers)

        if "amount" not in col_map:
            sys.exit(f"[ERROR] Could not find an 'amount' column. Detected headers: {headers}")

        rows = load_rows(ws, col_map)
        print(f"    Rows read: {len(rows)}")

        result = audit(rows)
        result["total_rows"] = len(rows)

        write_audit_sheet(wb, result, datetime.now())
        wb.save(local_path)
        print("[*] Audit sheet written to workbook.")

        print("[*] Uploading updated file to Google Drive…")
        upload_file(file_id, local_path)

    # ── console summary ───────────────────────────────────────────────────────
    print("\n" + "═" * 55)
    print("  AUDIT SUMMARY")
    print("═" * 55)
    print(f"  Rows read       : {result['total_rows']}")
    print(f"  Valid rows      : {result['valid_rows']}")
    print(f"  Grand total     : {result['grand_total']:,.2f}")
    print(f"  Categories      : {len(result['cat_totals'])}")
    print(f"  Anomalies ⚠️    : {len(result['anomalies'])}")
    print(f"  Duplicate groups: {len(result['duplicates'])}")
    if result["monthly"]:
        last_month, last_total = result["monthly"][-1]
        print(f"  Last month      : {last_month}  →  {last_total:,.2f}")
    print("═" * 55)
    print("  Top categories by spend:")
    for cat, total in sorted(result["cat_totals"].items(), key=lambda x: -x[1])[:5]:
        print(f"    {cat:<25} {total:>12,.2f}")
    print("═" * 55)
    print("  ✅ 'Audit Report' sheet added and file uploaded to Google Drive.")


if __name__ == "__main__":
    main()
